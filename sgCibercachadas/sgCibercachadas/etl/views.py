from django.shortcuts import render
from django.shortcuts import render,redirect
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from django.views import generic
from django.contrib import messages
from django.utils import timezone
from datetime import datetime,timezone
from etl.forms import RetornoForm, PotencialForm,ConsignaForm
from gerencial.models import *
from subprocess import Popen, PIPE
import openpyxl

class CargaDatos(LoginRequiredMixin,PermissionRequiredMixin,generic.TemplateView):
    template_name='etl/index.html'
    login_url='general:login'
    permission_required = 'gerencial.etl'

    def get(self,request,*args,**kwargs):
        return render(request, self.template_name)

class EtlBD(LoginRequiredMixin,PermissionRequiredMixin,generic.TemplateView):
    login_url='general:login'
    permission_required = 'gerencial.etl'

    def get(self,request,*args,**kwargs):
        return redirect('/etl')

    def post(self,request,*args,**kwargs):
        script_dir="../../ETL/etlSGcachadas"
        p = Popen(["python", "main.py"], cwd=script_dir, stdout=PIPE, stderr=PIPE)
        print(p.communicate())
        
        #bitacora 
        Bitacora.objects.create(
        usuario=request.user.first_name+" "+request.user.last_name,
        accion="Ejecucion ETL (Script) ",
        )

        messages.add_message(request, messages.SUCCESS, 'ETL ejecutado con exito')
        return redirect('/etl')

class CargaRetorno(LoginRequiredMixin,PermissionRequiredMixin,generic.TemplateView):
    login_url='general:login'
    permission_required = 'gerencial.etl'

    def get(self,request,*args,**kwargs):
        return redirect('/etl')
        
    def post(self,request,*args,**kwargs):
        form= RetornoForm(request.POST,request.FILES)
        if form.is_valid():             
            archivo=request.FILES.get("file_retorno")
            archivo= openpyxl.load_workbook(archivo)
            hoja1=archivo.get_sheet_by_name('Hoja1')
            max_row = hoja1.max_row
            for row in hoja1.iter_rows(min_row=5, max_col=6, max_row=max_row):
                i = 0
                retorno = ProductoRetorno()
                for cell in row:     
                    i += 1
                    if i==2:
                        if cell.value:
                            retorno.fecha = cell.value
                        else:
                            break
                    if i==3:
                        if cell.value:
                            try:
                                producto = Producto.objects.get(codigo=cell.value)
                                retorno.idProducto = producto
                                retorno.nombre_producto = producto.nombre
                                retorno.codigo = producto.codigo
                            except Producto.DoesNotExist:
                                retorno.nombre_producto = None
                        else:
                            break
                    if i==4:
                        if cell.value:
                            try:
                                cliente = Cliente.objects.get(nombre=cell.value) 
                                retorno.idCliente = cliente
                                retorno.nombre_cliente = cell.value 
                            except Cliente.DoesNotExist:
                                retorno.nombre_cliente = None
                        else:
                            break
                    if i==5:
                        if cell.value:
                            try:
                                proveedor = Proveedor.objects.get(razon_social=cell.value)
                                retorno.idProveedor = proveedor
                                nombre_proveedor = proveedor.razon_social
                            except Proveedor.DoesNotExist:
                                nombre_proveedor = None
                        else:
                            break
                    if i==6:
                        if cell.value:
                            retorno.cantidad = cell.value
                        else:
                            break
                if(retorno.nombre_producto and retorno.nombre_cliente and nombre_proveedor):
                    retorno.save()
                    #bitacora 
                    Bitacora.objects.create(
                    usuario=request.user.first_name+" "+request.user.last_name,
                    accion="Carga productos retorno (.xls) ",
                    )   
                      

            messages.add_message(request, messages.SUCCESS, 'Carga de producto retorno realizada con exito')
            return redirect('etl:carga_menu')
        else:
            return render(request,'etl/index.html',{'form':form})
            

class CargaPotencial(LoginRequiredMixin,PermissionRequiredMixin,generic.TemplateView):
    login_url='general:login'
    permission_required = 'gerencial.etl'

    def get(self,request,*args,**kwargs):
        return redirect('/etl')
    
    def post(self,request,*args,**kwargs):
        form= PotencialForm(request.POST,request.FILES)
        if form.is_valid():

            messages.add_message(request, messages.SUCCESS, 'Carga de producto potenciales realizada con exito')
            return redirect('etl:carga_menu')
        else:
            return render(request,'etl/index.html',{'form':form})


class CargaConsigna(LoginRequiredMixin,PermissionRequiredMixin,generic.TemplateView):
    login_url='general:login'
    permission_required = 'gerencial.etl'

    def get(self,request,*args,**kwargs):
        return redirect('/etl')
    
    def post(self,request,*args,**kwargs):
        form= ConsignaForm(request.POST,request.FILES)
        if form.is_valid():

            messages.add_message(request, messages.SUCCESS, 'Carga de productos en consigna realizada con exito')
            return redirect('etl:carga_menu')
        else:
            return render(request,'etl/index.html',{'form':form})


