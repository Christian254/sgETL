from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse
from datetime import datetime

def hoja_calculo(request,data,nombre,inicio,fin):
        wb = Workbook()
        ws = wb.active
        #En la celda A1 ponemos el texto 'REPORTE DE NOTAS'
        ws['A1'] = 'CiberCachada'
        #Combinamos las celdas
        ws.merge_cells('A1:D1')
        #este es el subtitulo
        ws['A2']= 'Clientes frecuentes'
        ws.merge_cells('A2:D2')
        #parametros utilizados
        ws['A3']= 'Periodo inicio: {} Periodo Fin: {}'.format(inicio,fin)
        ws.merge_cells('A3:E3')
        #Creamos los encabezados desde la celda B3 hasta la E
        ws['A5'] = 'Cliente'
        ws['B5'] = 'Cantidad Productos'
        ws['C5'] = 'Cantidad de Visitas'
        ws['D5'] = 'Ultima visita'
        #Nos posicionamos en la fila 4 para empezar a escribir
        cont=6
        #Recorremos el conjunto de notas
        #notas será la data que vamos a tener en el query
        for i in data:
            ws.cell(row=cont,column=1).value = i['idVenta__idCliente__nombre']
            ws.cell(row=cont,column=2).value = i['cantidad__sum']
            ws.cell(row=cont,column=3).value = i['idVenta__idCliente']
            ws.cell(row=cont,column=4).value = datetime.strftime(i['idVenta__fecha_hora'],"%d/%m/%Y")
            cont = cont + 1

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}.xlsx".format(nombre)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response